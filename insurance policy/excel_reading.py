import openpyxl

# Path to the Excel file
EXCEL_PATH = 'princeternship2025_premium_optimization_master_v3.xlsx'

def extract_excel_formulas(filepath=EXCEL_PATH):
    """
    Extracts and prints all formulas from every cell in every sheet of the Excel file.
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    formulas = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        formulas[sheet] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' or (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                    formulas[sheet][cell.coordinate] = cell.value
    return formulas

if __name__ == "__main__":
    formulas = extract_excel_formulas()
    for sheet, cells in formulas.items():
        print(f"Sheet: {sheet}")
        for coord, formula in cells.items():
            print(f"  {coord}: {formula}")
        print()